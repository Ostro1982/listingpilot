"""Bulk import from CSV/XLSX."""
import csv
import io
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_FIELDS = [
    "property_type", "operation", "address", "city", "state", "zip_code",
    "price", "bedrooms", "bathrooms", "living_area_sqft",
]
OPTIONAL_FIELDS = [
    "lot_size_sqft", "garage_spaces", "year_built", "hoa_fee", "features",
]
PHOTO_COLS = [f"photo{i}" for i in range(1, 9)]
ALL_COLS = REQUIRED_FIELDS + OPTIONAL_FIELDS + PHOTO_COLS


def _coerce(row: dict) -> tuple[dict, list[str]]:
    """Convert raw row to (property_dict, photos_list). Returns ({}, []) if invalid."""
    prop = {}
    for f in REQUIRED_FIELDS:
        v = (row.get(f) or "").strip() if isinstance(row.get(f), str) else row.get(f)
        if v in (None, "", []):
            raise ValueError(f"missing required field: {f}")
        prop[f] = v

    for f in OPTIONAL_FIELDS:
        v = row.get(f)
        if v in (None, ""):
            continue
        if isinstance(v, str):
            v = v.strip()
            if not v:
                continue
        prop[f] = v

    if "features" in prop and isinstance(prop["features"], str):
        prop["features"] = [x.strip() for x in prop["features"].split(",") if x.strip()]
    elif "features" not in prop:
        prop["features"] = []

    for k in ("price", "living_area_sqft", "lot_size_sqft", "hoa_fee"):
        if k in prop and prop[k] not in (None, ""):
            try:
                prop[k] = float(prop[k]) if k == "hoa_fee" else int(float(prop[k]))
            except (ValueError, TypeError):
                raise ValueError(f"invalid number in {k}: {prop[k]}")

    for k in ("bedrooms", "bathrooms"):
        prop[k] = str(prop[k])

    if "currency" not in prop:
        prop["currency"] = "USD"

    photos = []
    for col in PHOTO_COLS:
        v = row.get(col)
        if v and isinstance(v, str) and v.strip():
            photos.append(v.strip())

    return prop, photos


def parse_csv(file_obj) -> list[tuple[dict, list[str]]]:
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))
    rows = []
    for i, row in enumerate(reader, start=2):
        try:
            rows.append(_coerce(row))
        except ValueError as e:
            rows.append((None, [f"row {i}: {e}"]))
    return rows


def parse_xlsx(file_obj) -> list[tuple[dict, list[str]]]:
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    rows = []
    for i, row in enumerate(rows_iter, start=2):
        if not any(row):
            continue
        d = dict(zip(headers, row))
        try:
            rows.append(_coerce(d))
        except ValueError as e:
            rows.append((None, [f"row {i}: {e}"]))
    return rows


def parse_file(filename: str, file_obj) -> list[tuple[dict, list[str]]]:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return parse_csv(file_obj)
    if ext in (".xlsx", ".xlsm"):
        return parse_xlsx(file_obj)
    raise ValueError(f"unsupported file type: {ext}")


def template_csv_bytes() -> bytes:
    sample_rows = [
        {
            "property_type": "Single Family", "operation": "Sale",
            "address": "1234 Maple Ave", "city": "Austin", "state": "TX", "zip_code": "78704",
            "price": 495000, "bedrooms": 3, "bathrooms": 2, "living_area_sqft": 1850,
            "lot_size_sqft": 6500, "garage_spaces": 2, "year_built": 2018,
            "hoa_fee": 0, "features": "Pool, Hardwood floors, Updated kitchen",
            "photo1": "https://example.com/photo1.jpg",
            "photo2": "https://example.com/photo2.jpg",
            "photo3": "", "photo4": "", "photo5": "", "photo6": "", "photo7": "", "photo8": "",
        },
        {
            "property_type": "Condo", "operation": "Sale",
            "address": "450 Brickell Ave Apt 1208", "city": "Miami", "state": "FL", "zip_code": "33131",
            "price": 650000, "bedrooms": 2, "bathrooms": 2, "living_area_sqft": 1100,
            "lot_size_sqft": "", "garage_spaces": 1, "year_built": 2020,
            "hoa_fee": 850, "features": "Pool, Gym, Walk-in closet, Smart home",
            "photo1": "https://example.com/condo1.jpg",
            "photo2": "", "photo3": "", "photo4": "", "photo5": "", "photo6": "", "photo7": "", "photo8": "",
        },
    ]
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=ALL_COLS)
    w.writeheader()
    for r in sample_rows:
        w.writerow(r)
    return out.getvalue().encode("utf-8-sig")
